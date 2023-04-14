# -*- coding: utf-8 -*-
"""
Created on Sun Aug 21 14:34:38 2022

@author: Tony_lupara
"""
import pandas as pd
import itertools as it
import json
import collections.abc
import numpy as np
from openpyxl import load_workbook
import shutil
import logging

logging.basicConfig(filename=sys.argv[3], filemode='w')


mylines = []

with open (sys.argv[1], 'rt') as file:
    for line in file:
        line = line.splitlines()
        mylines.append(line)


genes = ["CYP2D6", "CYP2C19", "CYP2C9", "CYP2B6", "CYP1A2", "CYP3A4", "CYP3A5", "POR"]
script_path = "/analyze/soft/scripts/reports/pharm_stellar/libs/"

results = []
for line in mylines:
    for line in line:
        if line.startswith("*"):
            results.append(line)
        elif line.startswith("[*"):
            results.append(line[1:-1])

results = [x.replace('(full_gene_del)','D') for x in results]
results = [x.replace('*1/*6 or *4/*9','*1/*6') for x in results]
            
info_dict = dict(zip(genes, results))

##Create DataFrame for cyp2d6 alleles functions
genotype_cyp2d6_df = pd.read_csv(script_path + "CYP2D6_func_ref.txt", delimiter = "\t")
#genotype_cyp2d6_df = genotype_cyp2d6_df.drop(genotype_cyp2d6_df.columns[[2, 3, 4, 5, 6, 7, 8]], axis=1)
#genotype_cyp2d6_df = genotype_cyp2d6_df.drop([0])
#genotype_cyp2d6_df = genotype_cyp2d6_df.rename(columns={"GENE: CYP2D6": "Alleles", "Unnamed: 1": "Score"})
genotype_cyp2d6_dict = genotype_cyp2d6_df.set_index('Alleles').T.to_dict('list')

#val_dict = list(info_dict.values())

##Get cyp2d6 allles
val_cyp2d6_diplo = info_dict["CYP2D6"].split("/")
cyp2d6_diplo = '/'.join(val_cyp2d6_diplo)
val_cyp2d6_1 = val_cyp2d6_diplo[0]
val_cyp2d6_2 = val_cyp2d6_diplo[1]
#val_cyp2d6_1 = ['*22']


##Get cyp2d6 activity score
def return_score_cyp2d6(val):
    for geno, score in genotype_cyp2d6_dict.items():
        if geno==val:
            return score
    return('NA')

cyp2d6_allele1 = return_score_cyp2d6(val_cyp2d6_1)
cyp2d6_allele2 = return_score_cyp2d6(val_cyp2d6_2)


if cyp2d6_allele1=='NA':
    res1 = 'NA'
    print('ERROR: for CYP2D6 no match in database')
    logging.error('ERROR: for CYP2D6 no match in database')
elif 'Indeterminate' in cyp2d6_allele1:
    res1 = 'Indeterminate'
else:
    s1 = [str(integer) for integer in cyp2d6_allele1]
    a_string1 = "".join(s1)
    res1 = float(a_string1)

if cyp2d6_allele2=='NA':
    res2 = 'NA'
    print('ERROR: for CYP2D6 no match in database')
    logging.error('ERROR: for CYP2D6 no match in database')
elif 'Indeterminate' in cyp2d6_allele2:
    res2 = 'Indeterminate'
else:
    s2 = [str(integer) for integer in cyp2d6_allele2]
    a_string2 = "".join(s2)
    res2 = float(a_string2)


if 'NA' not in {res1, res2} and 'Indeterminate' not in {res1, res2}:
    cyp2d6_sum = res1 + res2
elif 'Indeterminate' in {res1, res2}:
    cyp2d6_sum = 'Indeterminate'
else:
    cyp2d6_sum = 'NA'

##Get cyp2d6 Phenotype
def return_phenotype_cyp2d6(val):
    if val=='Indeterminate':
        return('Indeterminate')
    elif val=='NA':
        return('NA')
    elif val==0:
        return('Poor Metabolizer')
    elif 1.25>val>=0.25:
        return('Intermediate Metabolizer')
    elif 2.25>=val>=1.25:
        return('Normal Metabolizer')
    elif val>2.25:
        return('Ultrarapid Metabolizer')
        
cyp2d6_pheno = return_phenotype_cyp2d6(cyp2d6_sum)
info_dict["CYP2D6"] = [info_dict["CYP2D6"]]
info_dict["CYP2D6"].append(cyp2d6_pheno)


##Create DataFrame for cyp2c19 alleles functions
genotype_cyp2c19_df = pd.read_csv(script_path + "CYP2C19_Dipl_Phen_Table.txt", delimiter = "\t")
#genotype_cyp2c19_df = genotype_cyp2c19_df.drop(genotype_cyp2c19_df.columns[[1, 3]], axis=1)
#genotype_cyp2c19_df = genotype_cyp2c19_df.rename(columns={"CYP2C19 Diplotype": "Alleles", "Coded Diplotype/Phenotype Summary": "Phenotype"})
genotype_cyp2c19_df_reverse = genotype_cyp2c19_df.copy()
genotype_cyp2c19_df_reverse['Alleles'] = genotype_cyp2c19_df.Alleles.str.split('/').apply(lambda x: '/'.join(x[::-1]))
genotype_cyp2c19_dict = genotype_cyp2c19_df.set_index('Alleles').T.to_dict('list')
genotype_cyp2c19_dict_reverse = genotype_cyp2c19_df_reverse.set_index('Alleles').T.to_dict('list')

##Get cyp2c19 Phenotype
def return_phenotype_cyp2c19(val):
    for geno, pheno in genotype_cyp2c19_dict.items():
        if geno==val:
            return pheno
    return('NA')

def return_phenotype_cyp2c19_reverse(val):
    for geno, pheno in genotype_cyp2c19_dict_reverse.items():
        if geno==val:
            return pheno
    return('NA')

cyp2c19_diplo = info_dict["CYP2C19"]
cyp2c19_pheno_forward = return_phenotype_cyp2c19(cyp2c19_diplo)
cyp2c19_pheno_forward = "".join(cyp2c19_pheno_forward)
cyp2c19_pheno_reverse = return_phenotype_cyp2c19_reverse(cyp2c19_diplo)
cyp2c19_pheno_reverse = "".join(cyp2c19_pheno_reverse)

cyp2c19_pheno = ""
if 'Indeterminate' in {cyp2c19_pheno_forward, cyp2c19_pheno_reverse}:
    cyp2c19_pheno = 'Indeterminate'
elif cyp2c19_pheno_forward != 'NA':
    cyp2c19_pheno = cyp2c19_pheno_forward
elif cyp2c19_pheno_reverse != 'NA':
    cyp2c19_pheno = cyp2c19_pheno_reverse
else:
    cyp2c19_pheno = 'NA'
    print('ERROR: for CYP2C19 no match in database')
    logging.error('ERROR: for CYP2C19 no match in database')

info_dict["CYP2C19"] = [info_dict["CYP2C19"]]
info_dict["CYP2C19"].append(cyp2c19_pheno)

##Create DataFrame for cyp2c9 alleles functions
genotype_cyp2c9_df = pd.read_csv(script_path + "CYP2C9_Dipl_Phen_Table.txt", delimiter = "\t")
#genotype_cyp2c9_df = genotype_cyp2c9_df.drop(genotype_cyp2c9_df.columns[[1, 3]], axis=1)
#genotype_cyp2c9_df = genotype_cyp2c9_df.rename(columns={"CYP2C9 Diplotype": "Alleles", "Coded Diplotype/Phenotype Summary": "Phenotype"})
genotype_cyp2c9_df_reverse = genotype_cyp2c9_df.copy()
genotype_cyp2c9_df_reverse['Alleles'] = genotype_cyp2c9_df.Alleles.str.split('/').apply(lambda x: '/'.join(x[::-1]))
genotype_cyp2c9_dict = genotype_cyp2c9_df.set_index('Alleles').T.to_dict('list')
genotype_cyp2c9_dict_reverse = genotype_cyp2c9_df_reverse.set_index('Alleles').T.to_dict('list')

##Get cyp2c9 Phenotype
def return_phenotype_cyp2c9(val):
    for geno, pheno in genotype_cyp2c9_dict.items():
        if geno==val:
            return pheno
    return('NA')

def return_phenotype_cyp2c9_reverse(val):
    for geno, pheno in genotype_cyp2c9_dict_reverse.items():
        if geno==val:
            return pheno
    return('NA')

cyp2c9_diplo = info_dict["CYP2C9"]
cyp2c9_pheno_forward = return_phenotype_cyp2c9(cyp2c9_diplo)
cyp2c9_pheno_forward = "".join(cyp2c9_pheno_forward)
cyp2c9_pheno_reverse = return_phenotype_cyp2c9_reverse(cyp2c9_diplo)
cyp2c9_pheno_reverse = "".join(cyp2c9_pheno_reverse)

cyp2c9_pheno = ""
if 'Indeterminate' in {cyp2c9_pheno_forward, cyp2c9_pheno_reverse}:
    cyp2c9_pheno = 'Indeterminate'
elif cyp2c9_pheno_forward != 'NA':
    cyp2c9_pheno = cyp2c9_pheno_forward
elif cyp2c9_pheno_reverse != 'NA':
    cyp2c9_pheno = cyp2c9_pheno_reverse
else:
    cyp2c9_pheno = 'NA'
    print('ERROR: for CYP2C9 no match in database')
    logging.error('ERROR: for CYP2C9 no match in database')

info_dict["CYP2C9"] = [info_dict["CYP2C9"]]
info_dict["CYP2C9"].append(cyp2c9_pheno)

##Create DataFrame for cyp2b6 alleles functions
genotype_cyp2b6_df = pd.read_csv(script_path + "CYP2B6_Dipl_Phen_Table.txt", delimiter = "\t")
#genotype_cyp2b6_df = genotype_cyp2b6_df.drop(genotype_cyp2b6_df.columns[[1, 3]], axis=1)
#genotype_cyp2b6_df = genotype_cyp2b6_df.rename(columns={"CYP2B6 Diplotype": "Alleles", "Coded Diplotype/Phenotype Summary": "Phenotype"})
genotype_cyp2b6_df_reverse = genotype_cyp2b6_df.copy()
genotype_cyp2b6_df_reverse['Alleles'] = genotype_cyp2b6_df.Alleles.str.split('/').apply(lambda x: '/'.join(x[::-1]))
genotype_cyp2b6_dict = genotype_cyp2b6_df.set_index('Alleles').T.to_dict('list')
genotype_cyp2b6_dict_reverse = genotype_cyp2b6_df_reverse.set_index('Alleles').T.to_dict('list')


##Get cyp2b6 Phenotype
def return_phenotype_cyp2b6(val):
    for geno, pheno in genotype_cyp2b6_dict.items():
        if geno==val:
            return pheno
    return('NA')
            

def return_phenotype_cyp2b6_reverse(val):
    for geno, pheno in genotype_cyp2b6_dict_reverse.items():
        if geno==val:
            return pheno
    return('NA')


cyp2b6_diplo = info_dict["CYP2B6"]
cyp2b6_pheno_forward = return_phenotype_cyp2b6(cyp2b6_diplo)
cyp2b6_pheno_forward = "".join(cyp2b6_pheno_forward)
cyp2b6_pheno_reverse = return_phenotype_cyp2b6_reverse(cyp2b6_diplo)
cyp2b6_pheno_reverse = "".join(cyp2b6_pheno_reverse)

cyp2b6_pheno = ""
if 'Indeterminate' in {cyp2b6_pheno_forward, cyp2b6_pheno_reverse}:
    cyp2b6_pheno = 'Indeterminate'
elif cyp2b6_pheno_forward != 'NA':
    cyp2b6_pheno = cyp2b6_pheno_forward
elif cyp2b6_pheno_reverse != 'NA':
    cyp2b6_pheno = cyp2b6_pheno_reverse
else:
    cyp2b6_pheno = 'NA'
    print('ERROR: for CYP2B6 no match in database')
    logging.error('ERROR: for CYP2B6 no match in database')


info_dict["CYP2B6"] = [info_dict["CYP2B6"]]
info_dict["CYP2B6"].append(cyp2b6_pheno)

##Create DataFrame for cyp3a5 alleles functions
genotype_cyp3a5_df = pd.read_csv(script_path + "CYP3A5_Dipl_Phen_Table.txt", delimiter = "\t")
#genotype_cyp3a5_df = genotype_cyp3a5_df.drop(genotype_cyp3a5_df.columns[[1, 3]], axis=1)
#genotype_cyp3a5_df = genotype_cyp3a5_df.rename(columns={"CYP3A5 Diplotype": "Alleles", "Coded Diplotype/Phenotype Summary": "Phenotype"})
genotype_cyp3a5_df_reverse = genotype_cyp3a5_df.copy()
genotype_cyp3a5_df_reverse['Alleles'] = genotype_cyp3a5_df.Alleles.str.split('/').apply(lambda x: '/'.join(x[::-1]))
genotype_cyp3a5_dict = genotype_cyp3a5_df.set_index('Alleles').T.to_dict('list')
genotype_cyp3a5_dict_reverse = genotype_cyp3a5_df_reverse.set_index('Alleles').T.to_dict('list')

##Get cyp3a5 Phenotype
def return_phenotype_cyp3a5(val):
    for geno, pheno in genotype_cyp3a5_dict.items():
        if geno==val:
            return pheno
    return('NA')

def return_phenotype_cyp3a5_reverse(val):
    for geno, pheno in genotype_cyp3a5_dict_reverse.items():
        if geno==val:
            return pheno
    return('NA')

cyp3a5_diplo = info_dict["CYP3A5"]
cyp3a5_pheno_forward = return_phenotype_cyp3a5(cyp3a5_diplo)
cyp3a5_pheno_forward = "".join(cyp3a5_pheno_forward)
cyp3a5_pheno_reverse = return_phenotype_cyp3a5_reverse(cyp3a5_diplo)
cyp3a5_pheno_reverse = "".join(cyp3a5_pheno_reverse)

cyp3a5_pheno = ""
if 'Indeterminate' in {cyp3a5_pheno_forward, cyp3a5_pheno_reverse}:
    cyp3a5_pheno = 'Indeterminate'
elif cyp3a5_pheno_forward != 'NA':
    cyp3a5_pheno = cyp3a5_pheno_forward
elif cyp3a5_pheno_reverse != 'NA':
    cyp3a5_pheno = cyp3a5_pheno_reverse
else:
    cyp3a5_pheno = 'NA'
    print('ERROR: for CYP3A5 no match in database')
    logging.error('ERROR: for CYP3A5 no match in database')

info_dict["CYP3A5"] = [info_dict["CYP3A5"]]
info_dict["CYP3A5"].append(cyp3a5_pheno)

##Create combinations for normal cyp1a2 alleles
alleles_norm = ['*1A', '*1B', '*1D', '*1E', '*1J', '*1L', '*1V', '*1W', '*1M']
norm_comb = [p for p in it.product(alleles_norm, repeat=2)]
norm_comb = list(map(str, norm_comb))
##Apply the same translation table to all strings
removetable = str.maketrans(' ', '/', "('),")
norm_comb = [s.translate(removetable) for s in norm_comb]
##dict for norm_met
norm_met_dict = { norm : "Normal Metabolizer" for norm in norm_comb }
norm_met_dict['*1C/*1F'] = "Normal Metabolizer"
norm_met_dict['*1F/*1C'] = "Normal Metabolizer"

##Create combinations for poor cyp1a2 alleles
alleles_poor_c = ['*1K', '*3', '*4', '*6', '*7', '*1C']
poor_comb = [p for p in it.product(alleles_poor_c, repeat=2)]
poor_comb = list(map(str, poor_comb))
##Apply the same translation table to all strings
removetable = str.maketrans(' ', '/', "('),")
poor_comb = [s.translate(removetable) for s in poor_comb]
##dict for poor_met
poor_comb_dict = { poor_met : "Poor Metabolizer" for poor_met in poor_comb }

##Get permutations of two lists to create intermediate alleles
alleles_norm_no_C = ['*1A', '*1B', '*1D', '*1E', '*1J', '*1L', '*1V', '*1W']
int_comb_isec_forward = [p for p in it.product(alleles_norm_no_C, alleles_poor_c, repeat=1)]
int_comb_isec_reverse = [p for p in it.product(alleles_poor_c, alleles_norm_no_C, repeat=1)]
int_comb_all = int_comb_isec_forward + int_comb_isec_reverse
int_comb_all = list(map(str, int_comb_all))
removetable = str.maketrans(' ', '/', "('),")
int_comb_all = [s.translate(removetable) for s in int_comb_all]
int_comb_dict = { int_met : "Intermediate Metabolizer" for int_met in int_comb_all }

##Create rapid alleles
alleles_rap = ['*1F']
rap_comb_forw = [p for p in it.product(alleles_norm_no_C, alleles_rap, repeat=1)]
rap_comb_rev = [p for p in it.product(alleles_rap, alleles_norm_no_C, repeat=1)]
rap_comb_all = rap_comb_forw + rap_comb_rev
rap_comb_all = list(map(str, rap_comb_all))
removetable = str.maketrans(' ', '/', "('),")
rap_comb_all = [s.translate(removetable) for s in rap_comb_all]
rap_comb_dict = { int_met : "Ultrarapid Metabolizer" for int_met in rap_comb_all }
rap_comb_dict['*1F/*1F'] = "Ultrarapid Metabolizer"

##concat dict
genotype_cyp1a2_dict = dict(norm_met_dict)
genotype_cyp1a2_dict.update(rap_comb_dict)
genotype_cyp1a2_dict.update(int_comb_dict)
genotype_cyp1a2_dict.update(poor_comb_dict)

##Get cyp1a2 Phenotype
def return_phenotype_cyp1a2(val):
    for geno, pheno in genotype_cyp1a2_dict.items():
        if geno==val:
            return pheno
    return('NA')

cyp1a2_diplo = info_dict["CYP1A2"]
cyp1a2_pheno = return_phenotype_cyp1a2(cyp1a2_diplo)

cyp1a2_pheno_result = ""
if cyp1a2_pheno=='NA':
    cyp1a2_pheno_result = 'NA'
    print('ERROR: for CYP1A2 no match in database')
    logging.error('ERROR: for CYP1A2 no match in database')
else:
    cyp1a2_pheno_result = cyp1a2_pheno

info_dict["CYP1A2"] = [info_dict["CYP1A2"]]
info_dict["CYP1A2"].append(cyp1a2_pheno_result)



##Create combinations for normal cyp3a4 alleles
alleles_norm_3a4 = ['*1', '*1G','*36']
norm_comb_3a4 = [p for p in it.product(alleles_norm_3a4, repeat=2)]
norm_comb_3a4 = list(map(str, norm_comb_3a4))
##Apply the same translation table to all strings
removetable = str.maketrans(' ', '/', "('),")
norm_comb_3a4 = [s.translate(removetable) for s in norm_comb_3a4]
##dict for norm_met
norm_met_dict_3a4 = { norm : "Normal Metabolizer" for norm in norm_comb_3a4 }

##Create combinations for poor cyp3a4 alleles
alleles_nofunc_3a4 = ['*6', '*20', '*26']
alleles_decreased_3a4 = ['*8', '*11', '*12', '*13', '*16', '*17', '*18', '*22']
nofunc_comb_3a4 = [p for p in it.product(alleles_nofunc_3a4, repeat=2)]
decreased_comb_3a4 = [p for p in it.product(alleles_decreased_3a4, repeat=2)]
poor_comb_3a4_forw = [p for p in it.product(alleles_nofunc_3a4, alleles_decreased_3a4, repeat=1)]
poor_comb_3a4_rev = [p for p in it.product(alleles_decreased_3a4, alleles_nofunc_3a4, repeat=1)]
nofunc_comb_3a4 = list(map(str, nofunc_comb_3a4))
decreased_comb_3a4 = list(map(str, decreased_comb_3a4))
poor_comb_3a4_forw = list(map(str, poor_comb_3a4_forw))
poor_comb_3a4_rev = list(map(str, poor_comb_3a4_rev))
##Apply the same translation table to all strings
removetable = str.maketrans(' ', '/', "('),")
nofunc_comb_3a4 = [s.translate(removetable) for s in nofunc_comb_3a4]
decreased_comb_3a4 = [s.translate(removetable) for s in decreased_comb_3a4]
poor_comb_3a4_forw = [s.translate(removetable) for s in poor_comb_3a4_forw]
poor_comb_3a4_rev = [s.translate(removetable) for s in poor_comb_3a4_rev]
poor_comb_all_3a4 = nofunc_comb_3a4 + decreased_comb_3a4 + poor_comb_3a4_forw + poor_comb_3a4_rev
##dict for poor_met
poor_met_dict_3a4 = { poor_met : "Poor Metabolizer" for poor_met in poor_comb_all_3a4 }

##Create combinations for intermediate cyp3a4 alleles
inter_comb_3a4_norm_decr_forw = [p for p in it.product(alleles_norm_3a4, alleles_decreased_3a4, repeat=1)]
inter_comb_3a4_norm_decr_rev = [p for p in it.product(alleles_decreased_3a4, alleles_norm_3a4, repeat=1)]
inter_comb_3a4_norm_nofunc_forw = [p for p in it.product(alleles_norm_3a4, alleles_nofunc_3a4, repeat=1)]
inter_comb_3a4_norm_nofunc_rev = [p for p in it.product(alleles_nofunc_3a4, alleles_norm_3a4, repeat=1)]
inter_comb_3a4_norm_decr_forw = list(map(str, inter_comb_3a4_norm_decr_forw))
inter_comb_3a4_norm_decr_rev = list(map(str, inter_comb_3a4_norm_decr_rev))
inter_comb_3a4_norm_nofunc_forw = list(map(str, inter_comb_3a4_norm_nofunc_forw))
inter_comb_3a4_norm_nofunc_rev = list(map(str, inter_comb_3a4_norm_nofunc_rev))
removetable = str.maketrans(' ', '/', "('),")
inter_comb_3a4_norm_decr_forw = [s.translate(removetable) for s in inter_comb_3a4_norm_decr_forw]
inter_comb_3a4_norm_decr_rev = [s.translate(removetable) for s in inter_comb_3a4_norm_decr_rev]
inter_comb_3a4_norm_nofunc_forw = [s.translate(removetable) for s in inter_comb_3a4_norm_nofunc_forw]
inter_comb_3a4_norm_nofunc_rev = [s.translate(removetable) for s in inter_comb_3a4_norm_nofunc_rev]
inter_comb_all_3a4 = inter_comb_3a4_norm_decr_forw + inter_comb_3a4_norm_decr_rev + inter_comb_3a4_norm_nofunc_forw + inter_comb_3a4_norm_nofunc_rev
inter_met_dict_3a4 = { inter_met : "Intermediate Metabolizer" for inter_met in inter_comb_all_3a4 }

##concat dict
genotype_cyp3a4_dict = dict(norm_met_dict_3a4)
genotype_cyp3a4_dict.update(poor_met_dict_3a4)
genotype_cyp3a4_dict.update(inter_met_dict_3a4)

##Get cyp3a4 Phenotype
def return_phenotype_cyp3a4(val):
    for geno, pheno in genotype_cyp3a4_dict.items():
        if geno==val:
            return pheno
    return('NA')

cyp3a4_diplo = info_dict["CYP3A4"]
cyp3a4_pheno = return_phenotype_cyp3a4(cyp3a4_diplo)

cyp3a4_pheno_result = ""
if cyp3a4_pheno=='NA':
    cyp3a4_pheno_result = 'NA'
    print('ERROR: for CYP3A4 no match in database')
    logging.error('ERROR: for CYP3A4 no match in database')
else:
    cyp3a4_pheno_result = cyp3a4_pheno

info_dict["CYP3A4"] = [info_dict["CYP3A4"]]
info_dict["CYP3A4"].append(cyp3a4_pheno_result)


##Get POR alleles
hap_por_diplo = info_dict["POR"].split("/")
por_diplo = '/'.join(hap_por_diplo)
hap_por_1 = hap_por_diplo[0]
hap_por_2 = hap_por_diplo[1]
por_genotype = hap_por_1 + '/' + hap_por_2


##Convert dict to df and write excel
info_dict.popitem()
result_df = pd.DataFrame(data=info_dict)
result_df = result_df.T.reset_index()
result_df = result_df.rename(columns={"index": "Ген", 0: "Аллели", 1: "Фенотип"})
#result_df.to_excel('H:/Spyder_work/Psycho_result.xlsx', index=False, sheet_name = 'Result')

#Change phenotype names to russian
result_df.loc[result_df['Фенотип'] == 'Normal Metabolizer', 'Фенотип'] = 'Нормальный метаболизатор'
result_df.loc[result_df['Фенотип'] == 'Intermediate Metabolizer', 'Фенотип'] = 'Промежуточный метаболизатор'
result_df.loc[result_df['Фенотип'] == 'Poor Metabolizer', 'Фенотип'] = 'Медленный метаболизатор'
result_df.loc[result_df['Фенотип'] == 'Ultrarapid Metabolizer', 'Фенотип'] = 'Ультрабыстрый метаболизатор'
result_df.loc[result_df['Фенотип'] == 'Rapid Metabolizer', 'Фенотип'] = 'Быстрый метаболизатор'
result_df.loc[result_df['Фенотип'] == 'Indeterminate', 'Фенотип'] = 'Неопределенный'


################################################
##Color Phenotype

def color_col(col, pattern_map, default=''):
    return np.select(
        [col.str.contains(k, na=False) for k in pattern_map.keys()],
        [f'background-color: {v}' for v in pattern_map.values()],
        default=default
    ).astype(str)

###############################################

##Conver dict to JSON
json_my = json.dumps(info_dict)

##Add risk alleles
#wb = load_workbook('C:/Users/acer/Desktop/Clinical_PGx/Risk_alleles.xlsx')
#writer = pd.ExcelWriter('C:/Users/acer/Desktop/Clinical_PGx/Risk_alleles.xlsx', engine = 'openpyxl')
#writer.book = wb
#result_df.to_excel(writer, index=False, sheet_name = 'Genotype')
#wb.move_sheet("Risk_alleles", (len(wb.sheetnames)-1))
#writer.save()
#writer.close()

##Load CPIC
cpic_data = pd.read_csv('C:/Users/acer/Desktop/Clinical_PGx/Drugs_from_exome/Drugs_PGx_from_exome_Twist.txt', delimiter = "\t", encoding='windows-1252')


###Drugs

##Neuroleptics
neuroleptics_table = pd.read_csv(script_path + "Neuroleptics.txt", delimiter = "\t")
neuroleptics_dict = neuroleptics_table.set_index('Drug').T.to_dict('list')

for gene, pheno in info_dict.items():
    for k, v in neuroleptics_dict.items():
        if gene==v[0]:
            val=info_dict.get(gene)
            neuroleptics_dict[k].append(val[1])

#Clozapine
Clozapine_dict = {'Clozapine': ['CYP1A2', 'CYP3A4', 'CYP2D6']}

clozapine_list = []
for gene, pheno in info_dict.items():
    for k, v in Clozapine_dict.items():
        if gene==v[0]:
            clozapine_list.append(pheno[1])
        elif gene==v[1]:
            clozapine_list.append(pheno[1])
        elif gene==v[2]:
            clozapine_list.append(pheno[1])

#def return_score(val):
#    for i in range(len(val)):
#        if val[i]=='Poor Metabolizer' or val[i]=='Likely Poor Metabolizer':
#            val[i]=int(0)
#        elif val[i]=='Intermediate Metabolizer' or val[i]=='Likely Intermediate Metabolizer':
#            val[i]=int(1)
#        elif val[i]=='Normal Metabolizer':
#            val[i]=int(2)
#        elif val[i]=='Ultrarapid Metabolizer' or val[i]=='Rapid Metabolizer':
#            val[i]=int(3)

metabo_to_int = {
    'Poor Metabolizer': int(0),
    'Likely Poor Metabolizer': int(0),
    'Intermediate Metabolizer': int(1),
    'Likely Intermediate Metabolizer': int(1),
    'Normal Metabolizer': int(2),
    'Ultrarapid Metabolizer': int(3),
    'Rapid Metabolizer': int(3),
    'Indeterminate': 'Indeterminate'
}

def return_score(val):
    for i in range(len(val)):
        param=str.strip(val[i])
        val[i]=metabo_to_int[param]

return_score(clozapine_list)

#for i in range(len(clozapine_list)):
#    if clozapine_list[i]=='Indeterminate':
#        clozapine_score = 'Indeterminate'
#    else:
#        clozapine_score = sum(clozapine_list)/len(clozapine_list)

if [el for el in clozapine_list if isinstance(el, collections.abc.Iterable) and ('Indeterminate' in el)]:
    clozapine_score = 'Indeterminate'
else:
    clozapine_score = sum(clozapine_list)/len(clozapine_list)
    
def return_phenotype(val):
    if val=='Indeterminate':
        return('Indeterminate')
    elif val==0:
        return('Poor Metabolizer')
    elif 1.25>val>=0.25:
        return('Intermediate Metabolizer')
    elif 2.25>=val>=1.25:
        return('Normal Metabolizer')
    elif val>2.25:
        return('Ultrarapid Metabolizer')
        
clozapine_pheno = return_phenotype(clozapine_score)
Clozapine_dict['Clozapine'].append(clozapine_pheno)


#Sertindole
Sertindole_dict = {'Sertindole': ['CYP3A4', 'CYP2D6']}

sertindole_list = []
for gene, pheno in info_dict.items():
    for k, v in Sertindole_dict.items():
        if gene==v[0]:
            sertindole_list.append(pheno[1])
        elif gene==v[1]:
            sertindole_list.append(pheno[1])

return_score(sertindole_list)

if [el for el in sertindole_list if isinstance(el, collections.abc.Iterable) and ('Indeterminate' in el)]:
    sertindole_score = 'Indeterminate'
else:
    sertindole_score = sum(sertindole_list)/len(sertindole_list)

sertindole_pheno = return_phenotype(sertindole_score)
Sertindole_dict['Sertindole'].append(sertindole_pheno)


#Convert Neuroleptics to df and write excel
neuroleptics_df = pd.DataFrame(data=neuroleptics_dict)
neuroleptics_df = neuroleptics_df.T.reset_index()
neuroleptics_df = neuroleptics_df.rename(columns={"index": "Drug", 0: "Gene(s)", 1: "Phenotype"})

clozapine_df = pd.DataFrame(data=Clozapine_dict)
clozapine_df = clozapine_df.T.reset_index()
clozapine_df['Gene(s)'] = clozapine_df[0] +", "+ clozapine_df[1] +", "+ clozapine_df[2]
clozapine_df = clozapine_df.drop(clozapine_df.columns[[1, 2, 3]], axis=1)
clozapine_df = clozapine_df.rename(columns={"index": "Drug", 3: "Phenotype"})
clozapine_df = clozapine_df[['Drug', 'Gene(s)', 'Phenotype']]
neuroleptics_df = pd.concat([neuroleptics_df, clozapine_df], ignore_index=True)

sertindole_df = pd.DataFrame(data=Sertindole_dict)
sertindole_df = sertindole_df.T.reset_index()
sertindole_df['Gene(s)'] = sertindole_df[0] +", "+ sertindole_df[1]
sertindole_df = sertindole_df.drop(sertindole_df.columns[[1, 2]], axis=1)
sertindole_df = sertindole_df.rename(columns={"index": "Drug", 2: "Phenotype"})
sertindole_df = sertindole_df[['Drug', 'Gene(s)', 'Phenotype']]
neuroleptics_df = pd.concat([neuroleptics_df, sertindole_df], ignore_index=True)

##Add alleles
neuroleptics_df.insert(2, 'Alleles', 'nan')
neuroleptics_df.loc[neuroleptics_df['Gene(s)'] == 'CYP2D6', 'Alleles'] = cyp2d6_diplo
neuroleptics_df.loc[neuroleptics_df['Gene(s)'] == 'CYP2B6', 'Alleles'] = cyp2b6_diplo
neuroleptics_df.loc[neuroleptics_df['Gene(s)'] == 'CYP2C9', 'Alleles'] = cyp2c9_diplo
neuroleptics_df.loc[neuroleptics_df['Gene(s)'] == 'CYP2C19', 'Alleles'] = cyp2c19_diplo
neuroleptics_df.loc[neuroleptics_df['Gene(s)'] == 'CYP1A2', 'Alleles'] = cyp1a2_diplo
neuroleptics_df.loc[neuroleptics_df['Gene(s)'] == 'CYP3A4', 'Alleles'] = cyp3a4_diplo
neuroleptics_df.loc[neuroleptics_df['Gene(s)'] == 'CYP3A5', 'Alleles'] = cyp3a5_diplo
clozapine_diplo = cyp1a2_diplo + ', ' + cyp3a4_diplo + ', ' + cyp2d6_diplo
neuroleptics_df.loc[neuroleptics_df['Gene(s)'] == 'CYP1A2, CYP3A4, CYP2D6', 'Alleles'] = clozapine_diplo
sertindole_diplo = cyp3a4_diplo + ', ' + cyp2d6_diplo
neuroleptics_df.loc[neuroleptics_df['Gene(s)'] == 'CYP3A4, CYP2D6', 'Alleles'] = sertindole_diplo


#neuroleptics_df.to_excel('H:/Spyder_work/Neuroleptics.xlsx', index=False)

#neuroleptics_colored = neuroleptics_df.style.apply(color_col,
#                                                   pattern_map={'Poor Metabolizer': 'red',
#                                                                'Intermediate Metabolizer': 'blue',
#                                                                'Normal Metabolizer': 'green',
#                                                                'Ultrarapid Metabolize': 'yellow',
#                                                               'Indeterminate': 'grey'},
#                                                  subset=['Phenotype'])

#Change phenotype names to russian
neuroleptics_df.loc[neuroleptics_df['Phenotype'] == 'Normal Metabolizer', 'Phenotype'] = 'Нормальный метаболизатор'
neuroleptics_df.loc[neuroleptics_df['Phenotype'] == 'Intermediate Metabolizer', 'Phenotype'] = 'Промежуточный метаболизатор'
neuroleptics_df.loc[neuroleptics_df['Phenotype'] == 'Poor Metabolizer', 'Phenotype'] = 'Медленный метаболизатор'
neuroleptics_df.loc[neuroleptics_df['Phenotype'] == 'Ultrarapid Metabolizer', 'Phenotype'] = 'Ультрабыстрый метаболизатор'
neuroleptics_df.loc[neuroleptics_df['Phenotype'] == 'Rapid Metabolizer', 'Phenotype'] = 'Быстрый метаболизатор'
neuroleptics_df.loc[neuroleptics_df['Phenotype'] == 'Indeterminate', 'Phenotype'] = 'Неопределенный'

#Change table names to russian
neuroleptics_df = neuroleptics_df.rename(columns={"Drug": "Препарат", "Gene(s)": "Ген", "Alleles": "Аллели", "Phenotype": "Фенотип"})

##Add recommendations and guidulines
neuroleptics_df.insert(4, 'Рекомендации по дозировке', 'nan')
neuroleptics_df.loc[neuroleptics_df['Фенотип'] == 'Нормальный метаболизатор', 'Рекомендации по дозировке'] = 'Использовать по инструкции'
neuroleptics_df.loc[neuroleptics_df['Фенотип'] == 'Медленный метаболизатор', 'Рекомендации по дозировке'] = 'Понизить дозировку'
neuroleptics_df.loc[neuroleptics_df['Фенотип'] == 'Промежуточный метаболизатор', 'Рекомендации по дозировке'] = 'Понизить дозировку'
neuroleptics_df.loc[neuroleptics_df['Фенотип'] == 'Быстрый метаболизатор', 'Рекомендации по дозировке'] = 'Повысить дозировку'
neuroleptics_df.loc[neuroleptics_df['Фенотип'] == 'Ультрабыстрый метаболизатор', 'Рекомендации по дозировке'] = 'Повысить дозировку'

neuroleptics_df_2 = neuroleptics_df.merge(cpic_data, left_on='Препарат',right_on='Drug', how='inner')
neuroleptics_df_3 = neuroleptics_df_2.drop(neuroleptics_df_2.columns[[5, 6, 7, 13, 14, 15]], axis=1)
neuroleptics_df_4 = neuroleptics_df_3.drop_duplicates(subset=['Препарат'])

neuroleptics_df_4.loc[neuroleptics_df_4['CPICLevelStatus'] == 'Provisional', 'CPICLevelStatus'] = 'Предварительный'
neuroleptics_df_4.loc[neuroleptics_df_4['CPICLevelStatus'] == 'Final', 'CPICLevelStatus'] = 'Окончательный'
neuroleptics_df_4.loc[neuroleptics_df_4['FDALabel'] == 'Actionable PGx', 'FDALabel'] = 'Активная'
neuroleptics_df_4.loc[neuroleptics_df_4['FDALabel'] == 'Informative PGx', 'FDALabel'] = 'Информирующая'
neuroleptics_df_4.loc[neuroleptics_df_4['FDALabel'] == 'Recommended genetic testing', 'FDALabel'] = 'Рекомендуется генетическое тестирование'
neuroleptics_df_4.loc[neuroleptics_df_4['FDALabel'] == 'Required genetic testing', 'FDALabel'] = 'Необходимо генетическое тестирование'
neuroleptics_df_5 = neuroleptics_df_4.rename(columns={"Guideline": "Методические рекомендации", "CPICLevel": "Уровень CPIC", "CPICLevelStatus": "Статус уровня CPIC", "PharmGKBLevel": "Уровень PharmGKB", "FDALabel": "Маркировка FDA"})


#Apply color
neuroleptics_colored = neuroleptics_df_5.style.apply(color_col,
                                                   pattern_map={'Медленный метаболизатор': '#FF0000',
                                                                'Промежуточный метаболизатор': '#FFFF00',
                                                                'Нормальный метаболизатор': '#00B050',
                                                                'Ультрабыстрый метаболизатор': '#00B0F0',
                                                                'Быстрый метаболизатор': '#15EBD2',
                                                                'Неопределенный': '#A6A6A6'},
                                                   subset=['Фенотип'])

#neuroleptics_colored.to_excel('C:/Users/acer/Desktop/Clinical_PGx/Neuroleptics.xlsx', sheet_name = 'Neuroleptics', engine='openpyxl', index=False)
#################################################


##Antidepressants
antidepressants_table = pd.read_csv(script_path + "Antidepressants.txt", delimiter = "\t")
antidepressants_dict = antidepressants_table.set_index('Drug').T.to_dict('list')

for gene, pheno in info_dict.items():
    for k, v in antidepressants_dict.items():
        if gene==v[0]:
            val=info_dict.get(gene)
            antidepressants_dict[k].append(val[1])

#Amitriptyline
Amitriptyline_dict = {'Amitriptyline': ['CYP2D6', 'CYP2C19']}

amitriptyline_list = []
for gene, pheno in info_dict.items():
    for k, v in Amitriptyline_dict.items():
        if gene==v[0]:
            amitriptyline_list.append(pheno[1])
        elif gene==v[1]:
            amitriptyline_list.append(pheno[1])

return_score(amitriptyline_list)

if [el for el in amitriptyline_list if isinstance(el, collections.abc.Iterable) and ('Indeterminate' in el)]:
    amitriptyline_score = 'Indeterminate'
else:
    amitriptyline_score = sum(amitriptyline_list)/len(amitriptyline_list)

amitriptyline_pheno = return_phenotype(amitriptyline_score)
Amitriptyline_dict['Amitriptyline'].append(amitriptyline_pheno)

#Clomipramine
Clomipramine_dict = {'Clomipramine': ['CYP2D6', 'CYP2C19']}

clomipramine_list = []
for gene, pheno in info_dict.items():
    for k, v in Clomipramine_dict.items():
        if gene==v[0]:
            clomipramine_list.append(pheno[1])
        elif gene==v[1]:
            clomipramine_list.append(pheno[1])

return_score(clomipramine_list)

if [el for el in clomipramine_list if isinstance(el, collections.abc.Iterable) and ('Indeterminate' in el)]:
    clomipramine_score = 'Indeterminate'
else:
    clomipramine_score = sum(clomipramine_list)/len(clomipramine_list)

clomipramine_pheno = return_phenotype(clomipramine_score)
Clomipramine_dict['Clomipramine'].append(clomipramine_pheno)

#Doxepine
Doxepine_dict = {'Doxepine': ['CYP2D6', 'CYP2C19']}

doxepine_list = []
for gene, pheno in info_dict.items():
    for k, v in Doxepine_dict.items():
        if gene==v[0]:
            doxepine_list.append(pheno[1])
        elif gene==v[1]:
            doxepine_list.append(pheno[1])

return_score(doxepine_list)

if [el for el in doxepine_list if isinstance(el, collections.abc.Iterable) and ('Indeterminate' in el)]:
    doxepine_score = 'Indeterminate'
else:
    doxepine_score = sum(doxepine_list)/len(doxepine_list)

doxepine_pheno = return_phenotype(doxepine_score)
Doxepine_dict['Doxepine'].append(doxepine_pheno)

#Imipramine
Imipramine_dict = {'Imipramine': ['CYP2D6', 'CYP2C19']}

imipramine_list = []
for gene, pheno in info_dict.items():
    for k, v in Imipramine_dict.items():
        if gene==v[0]:
            imipramine_list.append(pheno[1])
        elif gene==v[1]:
            imipramine_list.append(pheno[1])

return_score(imipramine_list)

if [el for el in imipramine_list if isinstance(el, collections.abc.Iterable) and ('Indeterminate' in el)]:
    imipramine_score = 'Indeterminate'
else:
    imipramine_score = sum(imipramine_list)/len(imipramine_list)

imipramine_pheno = return_phenotype(imipramine_score)
Imipramine_dict['Imipramine'].append(imipramine_pheno)

#S-Ketamin
S_Ketamin_dict = {'S-Ketamin': ['CYP2B6', 'CYP3A4']}

s_ketamin_list = []
for gene, pheno in info_dict.items():
    for k, v in S_Ketamin_dict.items():
        if gene==v[0]:
            s_ketamin_list.append(pheno[1])
        elif gene==v[1]:
            s_ketamin_list.append(pheno[1])


return_score(s_ketamin_list)

if [el for el in s_ketamin_list if isinstance(el, collections.abc.Iterable) and ('Indeterminate' in el)]:
    s_ketamin_score = 'Indeterminate'
else:
    s_ketamin_score = sum(s_ketamin_list)/len(s_ketamin_list)


s_ketamin_pheno = return_phenotype(s_ketamin_score)
S_Ketamin_dict['S-Ketamin'].append(s_ketamin_pheno)

#Fluoxetine
Fluoxetine_dict = {'Fluoxetine': ['CYP2D6', 'CYP2C9']}

fluoxetine_list = []
for gene, pheno in info_dict.items():
    for k, v in Fluoxetine_dict.items():
        if gene==v[0]:
            fluoxetine_list.append(pheno[1])
        elif gene==v[1]:
            fluoxetine_list.append(pheno[1])

return_score(fluoxetine_list)

if [el for el in fluoxetine_list if isinstance(el, collections.abc.Iterable) and ('Indeterminate' in el)]:
    fluoxetine_score = 'Indeterminate'
else:
    fluoxetine_score = sum(fluoxetine_list)/len(fluoxetine_list)

fluoxetine_pheno = return_phenotype(fluoxetine_score)
Fluoxetine_dict['Fluoxetine'].append(fluoxetine_pheno)

#Selegiline
Selegiline_dict = {'Selegiline': ['CYP2B6', 'CYP3A4']}

selegiline_list = []
for gene, pheno in info_dict.items():
    for k, v in Selegiline_dict.items():
        if gene==v[0]:
            selegiline_list.append(pheno[1])
        elif gene==v[1]:
            selegiline_list.append(pheno[1])

if selegiline_list[0]=='Poor Metabolizer':
    selegiline_list[0]=(int(0) * 0.75)
elif selegiline_list[0]=='Intermediate Metabolizer':
    selegiline_list[0]=(int(1) * 0.75)
elif selegiline_list[0]=='Normal Metabolizer':
    selegiline_list[0]=(int(2) * 0.75)
elif selegiline_list[0]=='Ultrarapid Metabolizer':
    selegiline_list[0]=(int(3) * 0.75)
elif selegiline_list[0]=='Rapid Metabolizer':
    selegiline_list[0]=(int(3) * 0.75)

if selegiline_list[1]=='Poor Metabolizer':
    selegiline_list[1]=(int(0) * 0.25)
elif selegiline_list[1]=='Intermediate Metabolizer':
    selegiline_list[1]=(int(1) * 0.25)
elif selegiline_list[1]=='Normal Metabolizer':
    selegiline_list[1]=(int(2) * 0.25)
elif selegiline_list[1]=='Ultrarapid Metabolizer':
    selegiline_list[1]=(int(3) * 0.25)

if [el for el in selegiline_list if isinstance(el, collections.abc.Iterable) and ('Indeterminate' in el)]:
    selegiline_score = 'Indeterminate'
else:
    selegiline_score = sum(selegiline_list)

selegiline_pheno = return_phenotype(selegiline_score)
Selegiline_dict['Selegiline'].append(selegiline_pheno)

#Convert Antidepressants to df and write excel
antidepressants_df = pd.DataFrame(data=antidepressants_dict)
antidepressants_df = antidepressants_df.T.reset_index()
antidepressants_df = antidepressants_df.rename(columns={"index": "Drug", 0: "Gene(s)", 1: "Phenotype"})

amitriptyline_df = pd.DataFrame(data=Amitriptyline_dict)
amitriptyline_df = amitriptyline_df.T.reset_index()
amitriptyline_df['Gene(s)'] = amitriptyline_df[0] +", "+ amitriptyline_df[1]
amitriptyline_df = amitriptyline_df.drop(amitriptyline_df.columns[[1, 2]], axis=1)
amitriptyline_df = amitriptyline_df.rename(columns={"index": "Drug", 2: "Phenotype"})
amitriptyline_df = amitriptyline_df[['Drug', 'Gene(s)', 'Phenotype']]
antidepressants_df = pd.concat([antidepressants_df, amitriptyline_df], ignore_index=True)

clomipramine_df = pd.DataFrame(data=Clomipramine_dict)
clomipramine_df = clomipramine_df.T.reset_index()
clomipramine_df['Gene(s)'] = clomipramine_df[0] +", "+ clomipramine_df[1]
clomipramine_df = clomipramine_df.drop(clomipramine_df.columns[[1, 2]], axis=1)
clomipramine_df = clomipramine_df.rename(columns={"index": "Drug", 2: "Phenotype"})
clomipramine_df = clomipramine_df[['Drug', 'Gene(s)', 'Phenotype']]
antidepressants_df = pd.concat([antidepressants_df, clomipramine_df], ignore_index=True)

doxepine_df = pd.DataFrame(data=Doxepine_dict)
doxepine_df = doxepine_df.T.reset_index()
doxepine_df['Gene(s)'] = doxepine_df[0] +", "+ doxepine_df[1]
doxepine_df = doxepine_df.drop(doxepine_df.columns[[1, 2]], axis=1)
doxepine_df = doxepine_df.rename(columns={"index": "Drug", 2: "Phenotype"})
doxepine_df = doxepine_df[['Drug', 'Gene(s)', 'Phenotype']]
antidepressants_df = pd.concat([antidepressants_df, doxepine_df], ignore_index=True)

imipramine_df = pd.DataFrame(data=Imipramine_dict)
imipramine_df = imipramine_df.T.reset_index()
imipramine_df['Gene(s)'] = imipramine_df[0] +", "+ imipramine_df[1]
imipramine_df = imipramine_df.drop(imipramine_df.columns[[1, 2]], axis=1)
imipramine_df = imipramine_df.rename(columns={"index": "Drug", 2: "Phenotype"})
imipramine_df = imipramine_df[['Drug', 'Gene(s)', 'Phenotype']]
antidepressants_df = pd.concat([antidepressants_df, imipramine_df], ignore_index=True)

s_ketamin_df = pd.DataFrame(data=S_Ketamin_dict)
s_ketamin_df = s_ketamin_df.T.reset_index()
s_ketamin_df['Gene(s)'] = s_ketamin_df[0] +", "+ s_ketamin_df[1]
s_ketamin_df = s_ketamin_df.drop(s_ketamin_df.columns[[1, 2]], axis=1)
s_ketamin_df = s_ketamin_df.rename(columns={"index": "Drug", 2: "Phenotype"})
s_ketamin_df = s_ketamin_df[['Drug', 'Gene(s)', 'Phenotype']]
antidepressants_df = pd.concat([antidepressants_df, s_ketamin_df], ignore_index=True)

fluoxetine_df = pd.DataFrame(data=Fluoxetine_dict)
fluoxetine_df = fluoxetine_df.T.reset_index()
fluoxetine_df['Gene(s)'] = fluoxetine_df[0] +", "+ fluoxetine_df[1]
fluoxetine_df = fluoxetine_df.drop(fluoxetine_df.columns[[1, 2]], axis=1)
fluoxetine_df = fluoxetine_df.rename(columns={"index": "Drug", 2: "Phenotype"})
fluoxetine_df = fluoxetine_df[['Drug', 'Gene(s)', 'Phenotype']]
antidepressants_df = pd.concat([antidepressants_df, fluoxetine_df], ignore_index=True)

selegiline_df = pd.DataFrame(data=Selegiline_dict)
selegiline_df = selegiline_df.T.reset_index()
selegiline_df['Gene(s)'] = selegiline_df[0] +", "+ selegiline_df[1]
selegiline_df = selegiline_df.drop(selegiline_df.columns[[1, 2]], axis=1)
selegiline_df = selegiline_df.rename(columns={"index": "Drug", 2: "Phenotype"})
selegiline_df = selegiline_df[['Drug', 'Gene(s)', 'Phenotype']]
antidepressants_df = pd.concat([antidepressants_df, selegiline_df], ignore_index=True)


##Add alleles
antidepressants_df.insert(2, 'Alleles', 'nan')
antidepressants_df.loc[antidepressants_df['Gene(s)'] == 'CYP2D6', 'Alleles'] = cyp2d6_diplo
antidepressants_df.loc[antidepressants_df['Gene(s)'] == 'CYP2B6', 'Alleles'] = cyp2b6_diplo
antidepressants_df.loc[antidepressants_df['Gene(s)'] == 'CYP2C9', 'Alleles'] = cyp2c9_diplo
antidepressants_df.loc[antidepressants_df['Gene(s)'] == 'CYP2C19', 'Alleles'] = cyp2c19_diplo
antidepressants_df.loc[antidepressants_df['Gene(s)'] == 'CYP1A2', 'Alleles'] = cyp1a2_diplo
antidepressants_df.loc[antidepressants_df['Gene(s)'] == 'CYP3A4', 'Alleles'] = cyp3a4_diplo
antidepressants_df.loc[antidepressants_df['Gene(s)'] == 'CYP3A5', 'Alleles'] = cyp3a5_diplo
amitriptyline_diplo = cyp2d6_diplo + ', ' + cyp2c19_diplo
antidepressants_df.loc[antidepressants_df['Gene(s)'] == 'CYP2D6, CYP2C19', 'Alleles'] = amitriptyline_diplo
clomipramine_diplo = cyp2d6_diplo + ', ' + cyp2c19_diplo
antidepressants_df.loc[antidepressants_df['Gene(s)'] == 'CYP2D6, CYP2C19', 'Alleles'] = clomipramine_diplo
doxepine_diplo = cyp2d6_diplo + ', ' + cyp2c19_diplo
antidepressants_df.loc[antidepressants_df['Gene(s)'] == 'CYP2D6, CYP2C19', 'Alleles'] = doxepine_diplo
imipramine_diplo = cyp2d6_diplo + ', ' + cyp2c19_diplo
antidepressants_df.loc[antidepressants_df['Gene(s)'] == 'CYP2D6, CYP2C19', 'Alleles'] = imipramine_diplo
s_ketamin_diplo = cyp2b6_diplo + ', ' + cyp3a4_diplo
antidepressants_df.loc[antidepressants_df['Gene(s)'] == 'CYP2B6, CYP3A4', 'Alleles'] = s_ketamin_diplo
fluoxetine_diplo = cyp2d6_diplo + ', ' + cyp2c9_diplo
antidepressants_df.loc[antidepressants_df['Gene(s)'] == 'CYP2D6, CYP2C9', 'Alleles'] = fluoxetine_diplo
selegiline_diplo = cyp2b6_diplo + ', ' + cyp3a4_diplo
antidepressants_df.loc[antidepressants_df['Gene(s)'] == 'CYP2B6, CYP3A4', 'Alleles'] = selegiline_diplo

#Change phenotype names to russian
antidepressants_df.loc[antidepressants_df['Phenotype'] == 'Normal Metabolizer', 'Phenotype'] = 'Нормальный метаболизатор'
antidepressants_df.loc[antidepressants_df['Phenotype'] == 'Intermediate Metabolizer', 'Phenotype'] = 'Промежуточный метаболизатор'
antidepressants_df.loc[antidepressants_df['Phenotype'] == 'Poor Metabolizer', 'Phenotype'] = 'Медленный метаболизатор'
antidepressants_df.loc[antidepressants_df['Phenotype'] == 'Ultrarapid Metabolizer', 'Phenotype'] = 'Ультрабыстрый метаболизатор'
antidepressants_df.loc[antidepressants_df['Phenotype'] == 'Rapid Metabolizer', 'Phenotype'] = 'Быстрый метаболизатор'
antidepressants_df.loc[antidepressants_df['Phenotype'] == 'Indeterminate', 'Phenotype'] = 'Неопределенный'

#Change table names to russian
antidepressants_df = antidepressants_df.rename(columns={"Drug": "Препарат", "Gene(s)": "Ген", "Alleles": "Аллели", "Phenotype": "Фенотип"})

##Add recommendations and guidulines
antidepressants_df.insert(4, 'Рекомендации по дозировке', 'nan')
antidepressants_df.loc[antidepressants_df['Фенотип'] == 'Нормальный метаболизатор', 'Рекомендации по дозировке'] = 'Использовать по инструкции'
antidepressants_df.loc[antidepressants_df['Фенотип'] == 'Медленный метаболизатор', 'Рекомендации по дозировке'] = 'Понизить дозировку'
antidepressants_df.loc[antidepressants_df['Фенотип'] == 'Промежуточный метаболизатор', 'Рекомендации по дозировке'] = 'Понизить дозировку'
antidepressants_df.loc[antidepressants_df['Фенотип'] == 'Быстрый метаболизатор', 'Рекомендации по дозировке'] = 'Повысить дозировку'
antidepressants_df.loc[antidepressants_df['Фенотип'] == 'Ультрабыстрый метаболизатор', 'Рекомендации по дозировке'] = 'Повысить дозировку'

antidepressants_df_2 = antidepressants_df.merge(cpic_data, left_on='Препарат',right_on='Drug', how='inner')
antidepressants_df_3 = antidepressants_df_2.drop(antidepressants_df_2.columns[[5, 6, 7, 13, 14, 15]], axis=1)
antidepressants_df_4 = antidepressants_df_3.drop_duplicates(subset=['Препарат'])

antidepressants_df_4.loc[antidepressants_df_4['CPICLevelStatus'] == 'Provisional', 'CPICLevelStatus'] = 'Предварительный'
antidepressants_df_4.loc[antidepressants_df_4['CPICLevelStatus'] == 'Final', 'CPICLevelStatus'] = 'Окончательный'
antidepressants_df_4.loc[antidepressants_df_4['FDALabel'] == 'Actionable PGx', 'FDALabel'] = 'Активная'
antidepressants_df_4.loc[antidepressants_df_4['FDALabel'] == 'Informative PGx', 'FDALabel'] = 'Информирующая'
antidepressants_df_4.loc[antidepressants_df_4['FDALabel'] == 'Recommended genetic testing', 'FDALabel'] = 'Рекомендуется генетическое тестирование'
antidepressants_df_4.loc[antidepressants_df_4['FDALabel'] == 'Required genetic testing', 'FDALabel'] = 'Необходимо генетическое тестирование'
antidepressants_df_5 = antidepressants_df_4.rename(columns={"Guideline": "Методические рекомендации", "CPICLevel": "Уровень CPIC", "CPICLevelStatus": "Статус уровня CPIC", "PharmGKBLevel": "Уровень PharmGKB", "FDALabel": "Маркировка FDA"})


#Apply color
antidepressants_colored = antidepressants_df_5.style.apply(color_col,
                                                   pattern_map={'Медленный метаболизатор': '#FF0000',
                                                                'Промежуточный метаболизатор': '#FFFF00',
                                                                'Нормальный метаболизатор': '#00B050',
                                                                'Ультрабыстрый метаболизатор': '#00B0F0',
                                                                'Быстрый метаболизатор': '#15EBD2',
                                                                'Неопределенный': '#A6A6A6'},
                                                   subset=['Фенотип'])
 
#antidepressants_colored.to_excel('H:/Spyder_work/Antidepressant.xlsx', engine='openpyxl', index=False)


##Anticonvulsants
anticonvulsants_table = pd.read_csv(script_path + "Anticonvulsants.txt", delimiter = "\t")
anticonvulsants_dict = anticonvulsants_table.set_index('Drug').T.to_dict('list')

for gene, pheno in info_dict.items():
    for k, v in anticonvulsants_dict.items():
        if gene==v[0]:
            val=info_dict.get(gene)
            anticonvulsants_dict[k].append(val[1])

#Diazepame
Diazepame_dict = {'Diazepame': ['CYP3A4', 'CYP2C19']}

diazepame_list = []
for gene, pheno in info_dict.items():
    for k, v in Diazepame_dict.items():
        if gene==v[0]:
            diazepame_list.append(pheno[1])
        elif gene==v[1]:
            diazepame_list.append(pheno[1])

return_score(diazepame_list)

if [el for el in diazepame_list if isinstance(el, collections.abc.Iterable) and ('Indeterminate' in el)]:
    diazepame_score = 'Indeterminate'
else:
    diazepame_score = sum(diazepame_list)/len(diazepame_list)

diazepame_pheno = return_phenotype(diazepame_score)
Diazepame_dict['Diazepame'].append(diazepame_pheno)

#Tiagabinum
Tiagabinum_dict = {'Tiagabinum': ['CYP3A4', 'CYP3A5']}

tiagabinum_list = []
for gene, pheno in info_dict.items():
    for k, v in Tiagabinum_dict.items():
        if gene==v[0]:
            tiagabinum_list.append(pheno[1])
        elif gene==v[1]:
            if v[1]=='CYP3A5' and pheno[1]=='Poor Metabolizer':
                tiagabinum_list.append('Normal Metabolizer')
            elif v[1]=='CYP3A5' and pheno[1]=='Intermediate Metabolizer':
                tiagabinum_list.append('Rapid Metabolizer')
            elif v[1]=='CYP3A5' and pheno[1]=='Normal Metabolizer':
                tiagabinum_list.append('Ultrarapid Metabolizer')

return_score(tiagabinum_list)

if [el for el in tiagabinum_list if isinstance(el, collections.abc.Iterable) and ('Indeterminate' in el)]:
    tiagabinum_score = 'Indeterminate'
else:
    tiagabinum_score = sum(tiagabinum_list)/len(tiagabinum_list)

tiagabinum_pheno = return_phenotype(tiagabinum_score)
Tiagabinum_dict['Tiagabinum'].append(tiagabinum_pheno)

#Convert Anticonvulsants to df and write excel
anticonvulsants_df = pd.DataFrame(data=anticonvulsants_dict)
anticonvulsants_df = anticonvulsants_df.T.reset_index()
anticonvulsants_df = anticonvulsants_df.rename(columns={"index": "Drug", 0: "Gene(s)", 1: "Phenotype"})

diazepame_df = pd.DataFrame(data=Diazepame_dict)
diazepame_df = diazepame_df.T.reset_index()
diazepame_df['Gene(s)'] = diazepame_df[0] +", "+ diazepame_df[1]
diazepame_df = diazepame_df.drop(diazepame_df.columns[[1, 2]], axis=1)
diazepame_df = diazepame_df.rename(columns={"index": "Drug", 2: "Phenotype"})
diazepame_df = diazepame_df[['Drug', 'Gene(s)', 'Phenotype']]
anticonvulsants_df = pd.concat([anticonvulsants_df, diazepame_df], ignore_index=True)

tiagabinum_df = pd.DataFrame(data=Tiagabinum_dict)
tiagabinum_df = tiagabinum_df.T.reset_index()
tiagabinum_df['Gene(s)'] = tiagabinum_df[0] +", "+ tiagabinum_df[1]
tiagabinum_df = tiagabinum_df.drop(tiagabinum_df.columns[[1, 2]], axis=1)
tiagabinum_df = tiagabinum_df.rename(columns={"index": "Drug", 2: "Phenotype"})
tiagabinum_df = tiagabinum_df[['Drug', 'Gene(s)', 'Phenotype']]
anticonvulsants_df = pd.concat([anticonvulsants_df, tiagabinum_df], ignore_index=True)

##Add alleles
anticonvulsants_df.insert(2, 'Alleles', 'nan')
anticonvulsants_df.loc[anticonvulsants_df['Gene(s)'] == 'CYP2D6', 'Alleles'] = cyp2d6_diplo
anticonvulsants_df.loc[anticonvulsants_df['Gene(s)'] == 'CYP2B6', 'Alleles'] = cyp2b6_diplo
anticonvulsants_df.loc[anticonvulsants_df['Gene(s)'] == 'CYP2C9', 'Alleles'] = cyp2c9_diplo
anticonvulsants_df.loc[anticonvulsants_df['Gene(s)'] == 'CYP2C19', 'Alleles'] = cyp2c19_diplo
anticonvulsants_df.loc[anticonvulsants_df['Gene(s)'] == 'CYP1A2', 'Alleles'] = cyp1a2_diplo
anticonvulsants_df.loc[anticonvulsants_df['Gene(s)'] == 'CYP3A4', 'Alleles'] = cyp3a4_diplo
anticonvulsants_df.loc[anticonvulsants_df['Gene(s)'] == 'CYP3A5', 'Alleles'] = cyp3a5_diplo
tiagabinum_diplo = cyp3a4_diplo + ', ' + cyp3a5_diplo
anticonvulsants_df.loc[anticonvulsants_df['Gene(s)'] == 'CYP3A4, CYP3A5', 'Alleles'] = tiagabinum_diplo
diazepame_diplo = cyp3a4_diplo + ', ' + cyp2c19_diplo
anticonvulsants_df.loc[anticonvulsants_df['Gene(s)'] == 'CYP3A4, CYP2C19', 'Alleles'] = diazepame_diplo

#Change phenotype names to russian
anticonvulsants_df.loc[anticonvulsants_df['Phenotype'] == 'Normal Metabolizer', 'Phenotype'] = 'Нормальный метаболизатор'
anticonvulsants_df.loc[anticonvulsants_df['Phenotype'] == 'Intermediate Metabolizer', 'Phenotype'] = 'Промежуточный метаболизатор'
anticonvulsants_df.loc[anticonvulsants_df['Phenotype'] == 'Poor Metabolizer', 'Phenotype'] = 'Медленный метаболизатор'
anticonvulsants_df.loc[anticonvulsants_df['Phenotype'] == 'Ultrarapid Metabolizer', 'Phenotype'] = 'Ультрабыстрый метаболизатор'
anticonvulsants_df.loc[anticonvulsants_df['Phenotype'] == 'Rapid Metabolizer', 'Phenotype'] = 'Быстрый метаболизатор'
anticonvulsants_df.loc[anticonvulsants_df['Phenotype'] == 'Indeterminate', 'Phenotype'] = 'Неопределенный'

#Change table names to russian
anticonvulsants_df = anticonvulsants_df.rename(columns={"Drug": "Препарат", "Gene(s)": "Ген", "Alleles": "Аллели", "Phenotype": "Фенотип"})

##Add recommendations and guidulines
anticonvulsants_df.insert(4, 'Рекомендации по дозировке', 'nan')
anticonvulsants_df.loc[anticonvulsants_df['Фенотип'] == 'Нормальный метаболизатор', 'Рекомендации по дозировке'] = 'Использовать по инструкции'
anticonvulsants_df.loc[anticonvulsants_df['Фенотип'] == 'Медленный метаболизатор', 'Рекомендации по дозировке'] = 'Понизить дозировку'
anticonvulsants_df.loc[anticonvulsants_df['Фенотип'] == 'Промежуточный метаболизатор', 'Рекомендации по дозировке'] = 'Понизить дозировку'
anticonvulsants_df.loc[anticonvulsants_df['Фенотип'] == 'Быстрый метаболизатор', 'Рекомендации по дозировке'] = 'Повысить дозировку'
anticonvulsants_df.loc[anticonvulsants_df['Фенотип'] == 'Ультрабыстрый метаболизатор', 'Рекомендации по дозировке'] = 'Повысить дозировку'

anticonvulsants_df_2 = anticonvulsants_df.merge(cpic_data, left_on='Препарат',right_on='Drug', how='inner')
anticonvulsants_df_3 = anticonvulsants_df_2.drop(anticonvulsants_df_2.columns[[5, 6, 7, 13, 14, 15]], axis=1)
anticonvulsants_df_4 = anticonvulsants_df_3.drop_duplicates(subset=['Препарат'])

anticonvulsants_df_4.loc[anticonvulsants_df_4['CPICLevelStatus'] == 'Provisional', 'CPICLevelStatus'] = 'Предварительный'
anticonvulsants_df_4.loc[anticonvulsants_df_4['CPICLevelStatus'] == 'Final', 'CPICLevelStatus'] = 'Окончательный'
anticonvulsants_df_4.loc[anticonvulsants_df_4['FDALabel'] == 'Actionable PGx', 'FDALabel'] = 'Активная'
anticonvulsants_df_4.loc[anticonvulsants_df_4['FDALabel'] == 'Informative PGx', 'FDALabel'] = 'Информирующая'
anticonvulsants_df_4.loc[anticonvulsants_df_4['FDALabel'] == 'Recommended genetic testing', 'FDALabel'] = 'Рекомендуется генетическое тестирование'
anticonvulsants_df_4.loc[anticonvulsants_df_4['FDALabel'] == 'Required genetic testing', 'FDALabel'] = 'Необходимо генетическое тестирование'
anticonvulsants_df_5 = anticonvulsants_df_4.rename(columns={"Guideline": "Методические рекомендации", "CPICLevel": "Уровень CPIC", "CPICLevelStatus": "Статус уровня CPIC", "PharmGKBLevel": "Уровень PharmGKB", "FDALabel": "Маркировка FDA"})


#Apply color
anticonvulsants_colored = anticonvulsants_df_5.style.apply(color_col,
                                                   pattern_map={'Медленный метаболизатор': '#FF0000',
                                                                'Промежуточный метаболизатор': '#FFFF00',
                                                                'Нормальный метаболизатор': '#00B050',
                                                                'Ультрабыстрый метаболизатор': '#00B0F0',
                                                                'Быстрый метаболизатор': '#15EBD2',
                                                                'Неопределенный': '#A6A6A6'},
                                                   subset=['Фенотип'])
 
#anticonvulsants_colored.to_excel('H:/Spyder_work/Anticonvulsants_Normotimics.xlsx', engine='openpyxl', index=False)


##Tranquilizers
tranquilizers_table = pd.read_csv(script_path + "Tranquilizers.txt", delimiter = "\t")
tranquilizers_dict = tranquilizers_table.set_index('Drug').T.to_dict('list')

for gene, pheno in info_dict.items():
    for k, v in tranquilizers_dict.items():
        if gene==v[0]:
            val=info_dict.get(gene)
            tranquilizers_dict[k].append(val[1])

#Clobazam
Clobazam_dict = {'Clobazam': ['CYP3A4', 'CYP2C19']}

clobazam_list = []
for gene, pheno in info_dict.items():
    for k, v in Clobazam_dict.items():
        if gene==v[0]:
            clobazam_list.append(pheno[1])
        elif gene==v[1]:
            clobazam_list.append(pheno[1])

return_score(clobazam_list)

if [el for el in clobazam_list if isinstance(el, collections.abc.Iterable) and ('Indeterminate' in el)]:
    clobazam_score = 'Indeterminate'
else:
    clobazam_score = sum(clobazam_list)/len(clobazam_list)

clobazam_pheno = return_phenotype(clobazam_score)
Clobazam_dict['Clobazam'].append(clobazam_pheno)

#Hydroxyzine
Hydroxyzine_dict = {'Hydroxyzine': ['CYP3A4', 'CYP2C19']}

hydroxyzine_list = []
for gene, pheno in info_dict.items():
    for k, v in Hydroxyzine_dict.items():
        if gene==v[0]:
            hydroxyzine_list.append(pheno[1])
        elif gene==v[1]:
            hydroxyzine_list.append(pheno[1])

return_score(hydroxyzine_list)

if [el for el in hydroxyzine_list if isinstance(el, collections.abc.Iterable) and ('Indeterminate' in el)]:
    hydroxyzine_score = 'Indeterminate'
else:
    hydroxyzine_score = sum(hydroxyzine_list)/len(hydroxyzine_list)

hydroxyzine_pheno = return_phenotype(hydroxyzine_score)
Hydroxyzine_dict['Hydroxyzine'].append(hydroxyzine_pheno)

#Alprazolam
Alprazolam_dict = {'Alprazolam': ['CYP3A4', 'CYP3A5']}

alprazolam_list = []
for gene, pheno in info_dict.items():
    for k, v in Alprazolam_dict.items():
        if gene==v[0]:
            alprazolam_list.append(pheno[1])
        elif gene==v[1]:
            if v[1]=='CYP3A5' and pheno[1]=='Poor Metabolizer':
                alprazolam_list.append('Normal Metabolizer')
            elif v[1]=='CYP3A5' and pheno[1]=='Intermediate Metabolizer':
                alprazolam_list.append('Rapid Metabolizer')
            elif v[1]=='CYP3A5' and pheno[1]=='Normal Metabolizer':
                alprazolam_list.append('Ultrarapid Metabolizer')

return_score(alprazolam_list)

if [el for el in alprazolam_list if isinstance(el, collections.abc.Iterable) and ('Indeterminate' in el)]:
    alprazolam_score = 'Indeterminate'
else:
    alprazolam_score = sum(alprazolam_list)/len(alprazolam_list)

alprazolam_pheno = return_phenotype(alprazolam_score)
Alprazolam_dict['Alprazolam'].append(alprazolam_pheno)

#Triazolam
Triazolam_dict = {'Triazolam': ['CYP3A4', 'CYP3A5']}

triazolam_list = []
for gene, pheno in info_dict.items():
    for k, v in Triazolam_dict.items():
        if gene==v[0]:
            triazolam_list.append(pheno[1])
        elif gene==v[1]:
            if v[1]=='CYP3A5' and pheno[1]=='Poor Metabolizer':
                triazolam_list.append('Normal Metabolizer')
            elif v[1]=='CYP3A5' and pheno[1]=='Intermediate Metabolizer':
                triazolam_list.append('Rapid Metabolizer')
            elif v[1]=='CYP3A5' and pheno[1]=='Normal Metabolizer':
                triazolam_list.append('Ultrarapid Metabolizer')

return_score(triazolam_list)

if [el for el in triazolam_list if isinstance(el, collections.abc.Iterable) and ('Indeterminate' in el)]:
    triazolam_score = 'Indeterminate'
else:
    triazolam_score = sum(triazolam_list)/len(triazolam_list)

triazolam_pheno = return_phenotype(triazolam_score)
Triazolam_dict['Triazolam'].append(triazolam_pheno)

#Convert Tranquilizers to df and write excel
tranquilizers_df = pd.DataFrame(data=tranquilizers_dict)
tranquilizers_df = tranquilizers_df.T.reset_index()
tranquilizers_df = tranquilizers_df.rename(columns={"index": "Drug", 0: "Gene(s)", 1: "Phenotype"})

clobazam_df = pd.DataFrame(data=Clobazam_dict)
clobazam_df = clobazam_df.T.reset_index()
clobazam_df['Gene(s)'] = clobazam_df[0] +", "+ clobazam_df[1]
clobazam_df = clobazam_df.drop(clobazam_df.columns[[1, 2]], axis=1)
clobazam_df = clobazam_df.rename(columns={"index": "Drug", 2: "Phenotype"})
clobazam_df = clobazam_df[['Drug', 'Gene(s)', 'Phenotype']]
tranquilizers_df = pd.concat([tranquilizers_df, clobazam_df], ignore_index=True)

hydroxyzine_df = pd.DataFrame(data=Hydroxyzine_dict)
hydroxyzine_df = hydroxyzine_df.T.reset_index()
hydroxyzine_df['Gene(s)'] = hydroxyzine_df[0] +", "+ hydroxyzine_df[1]
hydroxyzine_df = hydroxyzine_df.drop(hydroxyzine_df.columns[[1, 2]], axis=1)
hydroxyzine_df = hydroxyzine_df.rename(columns={"index": "Drug", 2: "Phenotype"})
hydroxyzine_df = hydroxyzine_df[['Drug', 'Gene(s)', 'Phenotype']]
tranquilizers_df = pd.concat([tranquilizers_df, hydroxyzine_df], ignore_index=True)

alprazolam_df = pd.DataFrame(data=Alprazolam_dict)
alprazolam_df = alprazolam_df.T.reset_index()
alprazolam_df['Gene(s)'] = alprazolam_df[0] +", "+ alprazolam_df[1]
alprazolam_df = alprazolam_df.drop(alprazolam_df.columns[[1, 2]], axis=1)
alprazolam_df = alprazolam_df.rename(columns={"index": "Drug", 2: "Phenotype"})
alprazolam_df = alprazolam_df[['Drug', 'Gene(s)', 'Phenotype']]
tranquilizers_df = pd.concat([tranquilizers_df, alprazolam_df], ignore_index=True)

triazolam_df = pd.DataFrame(data=Triazolam_dict)
triazolam_df = triazolam_df.T.reset_index()
triazolam_df['Gene(s)'] = triazolam_df[0] +", "+ triazolam_df[1]
triazolam_df = triazolam_df.drop(triazolam_df.columns[[1, 2]], axis=1)
triazolam_df = triazolam_df.rename(columns={"index": "Drug", 2: "Phenotype"})
triazolam_df = triazolam_df[['Drug', 'Gene(s)', 'Phenotype']]
tranquilizers_df = pd.concat([tranquilizers_df, triazolam_df], ignore_index=True)

##Add alleles
tranquilizers_df.insert(2, 'Alleles', 'nan')
tranquilizers_df.loc[tranquilizers_df['Gene(s)'] == 'CYP2D6', 'Alleles'] = cyp2d6_diplo
tranquilizers_df.loc[tranquilizers_df['Gene(s)'] == 'CYP2B6', 'Alleles'] = cyp2b6_diplo
tranquilizers_df.loc[tranquilizers_df['Gene(s)'] == 'CYP2C9', 'Alleles'] = cyp2c9_diplo
tranquilizers_df.loc[tranquilizers_df['Gene(s)'] == 'CYP2C19', 'Alleles'] = cyp2c19_diplo
tranquilizers_df.loc[tranquilizers_df['Gene(s)'] == 'CYP1A2', 'Alleles'] = cyp1a2_diplo
tranquilizers_df.loc[tranquilizers_df['Gene(s)'] == 'CYP3A4', 'Alleles'] = cyp3a4_diplo
tranquilizers_df.loc[tranquilizers_df['Gene(s)'] == 'CYP3A5', 'Alleles'] = cyp3a5_diplo
clobazam_diplo = cyp3a4_diplo + ', ' + cyp2c19_diplo
tranquilizers_df.loc[tranquilizers_df['Gene(s)'] == 'CYP3A4, CYP2C19', 'Alleles'] = clobazam_diplo
hydroxyzine_diplo = cyp3a4_diplo + ', ' + cyp2c19_diplo
tranquilizers_df.loc[tranquilizers_df['Gene(s)'] == 'CYP3A4, CYP2C19', 'Alleles'] = hydroxyzine_diplo
alprazolam_diplo = cyp3a4_diplo + ', ' + cyp3a5_diplo
tranquilizers_df.loc[tranquilizers_df['Gene(s)'] == 'CYP3A4, CYP3A5', 'Alleles'] = alprazolam_diplo
triazolam_diplo = cyp3a4_diplo + ', ' + cyp3a5_diplo
tranquilizers_df.loc[tranquilizers_df['Gene(s)'] == 'CYP3A4, CYP3A5', 'Alleles'] = triazolam_diplo

#Change phenotype names to russian
tranquilizers_df.loc[tranquilizers_df['Phenotype'] == 'Normal Metabolizer', 'Phenotype'] = 'Нормальный метаболизатор'
tranquilizers_df.loc[tranquilizers_df['Phenotype'] == 'Intermediate Metabolizer', 'Phenotype'] = 'Промежуточный метаболизатор'
tranquilizers_df.loc[tranquilizers_df['Phenotype'] == 'Poor Metabolizer', 'Phenotype'] = 'Медленный метаболизатор'
tranquilizers_df.loc[tranquilizers_df['Phenotype'] == 'Ultrarapid Metabolizer', 'Phenotype'] = 'Ультрабыстрый метаболизатор'
tranquilizers_df.loc[tranquilizers_df['Phenotype'] == 'Rapid Metabolizer', 'Phenotype'] = 'Быстрый метаболизатор'
tranquilizers_df.loc[tranquilizers_df['Phenotype'] == 'Indeterminate', 'Phenotype'] = 'Неопределенный'

#Change table names to russian
tranquilizers_df = tranquilizers_df.rename(columns={"Drug": "Препарат", "Gene(s)": "Ген", "Alleles": "Аллели", "Phenotype": "Фенотип"})

##Add recommendations and guidulines
tranquilizers_df.insert(4, 'Рекомендации по дозировке', 'nan')
tranquilizers_df.loc[tranquilizers_df['Фенотип'] == 'Нормальный метаболизатор', 'Рекомендации по дозировке'] = 'Использовать по инструкции'
tranquilizers_df.loc[tranquilizers_df['Фенотип'] == 'Медленный метаболизатор', 'Рекомендации по дозировке'] = 'Понизить дозировку'
tranquilizers_df.loc[tranquilizers_df['Фенотип'] == 'Промежуточный метаболизатор', 'Рекомендации по дозировке'] = 'Понизить дозировку'
tranquilizers_df.loc[tranquilizers_df['Фенотип'] == 'Быстрый метаболизатор', 'Рекомендации по дозировке'] = 'Повысить дозировку'
tranquilizers_df.loc[tranquilizers_df['Фенотип'] == 'Ультрабыстрый метаболизатор', 'Рекомендации по дозировке'] = 'Повысить дозировку'

tranquilizers_df_2 = tranquilizers_df.merge(cpic_data, left_on='Препарат',right_on='Drug', how='inner')
tranquilizers_df_3 = tranquilizers_df_2.drop(tranquilizers_df_2.columns[[5, 6, 7, 13, 14, 15]], axis=1)
tranquilizers_df_4 = tranquilizers_df_3.drop_duplicates(subset=['Препарат'])

tranquilizers_df_4.loc[tranquilizers_df_4['CPICLevelStatus'] == 'Provisional', 'CPICLevelStatus'] = 'Предварительный'
tranquilizers_df_4.loc[tranquilizers_df_4['CPICLevelStatus'] == 'Final', 'CPICLevelStatus'] = 'Окончательный'
tranquilizers_df_4.loc[tranquilizers_df_4['FDALabel'] == 'Actionable PGx', 'FDALabel'] = 'Активная'
tranquilizers_df_4.loc[tranquilizers_df_4['FDALabel'] == 'Informative PGx', 'FDALabel'] = 'Информирующая'
tranquilizers_df_4.loc[tranquilizers_df_4['FDALabel'] == 'Recommended genetic testing', 'FDALabel'] = 'Рекомендуется генетическое тестирование'
tranquilizers_df_4.loc[tranquilizers_df_4['FDALabel'] == 'Required genetic testing', 'FDALabel'] = 'Необходимо генетическое тестирование'
tranquilizers_df_5 = tranquilizers_df_4.rename(columns={"Guideline": "Методические рекомендации", "CPICLevel": "Уровень CPIC", "CPICLevelStatus": "Статус уровня CPIC", "PharmGKBLevel": "Уровень PharmGKB", "FDALabel": "Маркировка FDA"})


#Apply color
tranquilizers_colored = tranquilizers_df_5.style.apply(color_col,
                                                   pattern_map={'Медленный метаболизатор': '#FF0000',
                                                                'Промежуточный метаболизатор': '#FFFF00',
                                                                'Нормальный метаболизатор': '#00B050',
                                                                'Ультрабыстрый метаболизатор': '#00B0F0',
                                                                'Быстрый метаболизатор': '#15EBD2',
                                                                'Неопределенный': '#A6A6A6'},
                                                   subset=['Фенотип'])
 
#tranquilizers_colored.to_excel('H:/Spyder_work/Tranquilizers_Sedatives.xlsx', engine='openpyxl', index=False)


##Nootropic, Psychostimulants
nootropic_stimulants_table = pd.read_csv(script_path + "Nootropic_Stimulants.txt", delimiter = "\t")
nootropic_stimulants_dict = nootropic_stimulants_table.set_index('Drug').T.to_dict('list')

for gene, pheno in info_dict.items():
    for k, v in nootropic_stimulants_dict.items():
        if gene==v[0]:
            val=info_dict.get(gene)
            nootropic_stimulants_dict[k].append(val[1])

#Convert Nootropic, Psychostimulants to df and write excel
nootropic_stimulants_df = pd.DataFrame(data=nootropic_stimulants_dict)
nootropic_stimulants_df = nootropic_stimulants_df.T.reset_index()
nootropic_stimulants_df = nootropic_stimulants_df.rename(columns={"index": "Drug", 0: "Gene(s)", 1: "Phenotype"})

##Add alleles
nootropic_stimulants_df.insert(2, 'Alleles', 'nan')
nootropic_stimulants_df.loc[nootropic_stimulants_df['Gene(s)'] == 'CYP2D6', 'Alleles'] = cyp2d6_diplo
nootropic_stimulants_df.loc[nootropic_stimulants_df['Gene(s)'] == 'CYP2B6', 'Alleles'] = cyp2b6_diplo
nootropic_stimulants_df.loc[nootropic_stimulants_df['Gene(s)'] == 'CYP2C9', 'Alleles'] = cyp2c9_diplo
nootropic_stimulants_df.loc[nootropic_stimulants_df['Gene(s)'] == 'CYP2C19', 'Alleles'] = cyp2c19_diplo
nootropic_stimulants_df.loc[nootropic_stimulants_df['Gene(s)'] == 'CYP1A2', 'Alleles'] = cyp1a2_diplo
nootropic_stimulants_df.loc[nootropic_stimulants_df['Gene(s)'] == 'CYP3A4', 'Alleles'] = cyp3a4_diplo
nootropic_stimulants_df.loc[nootropic_stimulants_df['Gene(s)'] == 'CYP3A5', 'Alleles'] = cyp3a5_diplo

#Change phenotype names to russian
nootropic_stimulants_df.loc[nootropic_stimulants_df['Phenotype'] == 'Normal Metabolizer', 'Phenotype'] = 'Нормальный метаболизатор'
nootropic_stimulants_df.loc[nootropic_stimulants_df['Phenotype'] == 'Intermediate Metabolizer', 'Phenotype'] = 'Промежуточный метаболизатор'
nootropic_stimulants_df.loc[nootropic_stimulants_df['Phenotype'] == 'Poor Metabolizer', 'Phenotype'] = 'Медленный метаболизатор'
nootropic_stimulants_df.loc[nootropic_stimulants_df['Phenotype'] == 'Ultrarapid Metabolizer', 'Phenotype'] = 'Ультрабыстрый метаболизатор'
nootropic_stimulants_df.loc[nootropic_stimulants_df['Phenotype'] == 'Rapid Metabolizer', 'Phenotype'] = 'Быстрый метаболизатор'
nootropic_stimulants_df.loc[nootropic_stimulants_df['Phenotype'] == 'Indeterminate', 'Phenotype'] = 'Неопределенный'

#Change table names to russian
nootropic_stimulants_df = nootropic_stimulants_df.rename(columns={"Drug": "Препарат", "Gene(s)": "Ген", "Alleles": "Аллели", "Phenotype": "Фенотип"})

##Add recommendations and guidulines
nootropic_stimulants_df.insert(4, 'Рекомендации по дозировке', 'nan')
nootropic_stimulants_df.loc[nootropic_stimulants_df['Фенотип'] == 'Нормальный метаболизатор', 'Рекомендации по дозировке'] = 'Использовать по инструкции'
nootropic_stimulants_df.loc[nootropic_stimulants_df['Фенотип'] == 'Медленный метаболизатор', 'Рекомендации по дозировке'] = 'Понизить дозировку'
nootropic_stimulants_df.loc[nootropic_stimulants_df['Фенотип'] == 'Промежуточный метаболизатор', 'Рекомендации по дозировке'] = 'Понизить дозировку'
nootropic_stimulants_df.loc[nootropic_stimulants_df['Фенотип'] == 'Быстрый метаболизатор', 'Рекомендации по дозировке'] = 'Повысить дозировку'
nootropic_stimulants_df.loc[nootropic_stimulants_df['Фенотип'] == 'Ультрабыстрый метаболизатор', 'Рекомендации по дозировке'] = 'Повысить дозировку'

nootropic_stimulants_df_2 = nootropic_stimulants_df.merge(cpic_data, left_on='Препарат',right_on='Drug', how='inner')
nootropic_stimulants_df_3 = nootropic_stimulants_df_2.drop(nootropic_stimulants_df_2.columns[[5, 6, 7, 13, 14, 15]], axis=1)
nootropic_stimulants_df_4 = nootropic_stimulants_df_3.drop_duplicates(subset=['Препарат'])

nootropic_stimulants_df_4.loc[nootropic_stimulants_df_4['CPICLevelStatus'] == 'Provisional', 'CPICLevelStatus'] = 'Предварительный'
nootropic_stimulants_df_4.loc[nootropic_stimulants_df_4['CPICLevelStatus'] == 'Final', 'CPICLevelStatus'] = 'Окончательный'
nootropic_stimulants_df_4.loc[nootropic_stimulants_df_4['FDALabel'] == 'Actionable PGx', 'FDALabel'] = 'Активная'
nootropic_stimulants_df_4.loc[nootropic_stimulants_df_4['FDALabel'] == 'Informative PGx', 'FDALabel'] = 'Информирующая'
nootropic_stimulants_df_4.loc[nootropic_stimulants_df_4['FDALabel'] == 'Recommended genetic testing', 'FDALabel'] = 'Рекомендуется генетическое тестирование'
nootropic_stimulants_df_4.loc[nootropic_stimulants_df_4['FDALabel'] == 'Required genetic testing', 'FDALabel'] = 'Необходимо генетическое тестирование'
nootropic_stimulants_df_5 = nootropic_stimulants_df_4.rename(columns={"Guideline": "Методические рекомендации", "CPICLevel": "Уровень CPIC", "CPICLevelStatus": "Статус уровня CPIC", "PharmGKBLevel": "Уровень PharmGKB", "FDALabel": "Маркировка FDA"})


#Apply color
nootropic_stimulants_colored = nootropic_stimulants_df_5.style.apply(color_col,
                                                   pattern_map={'Медленный метаболизатор': '#FF0000',
                                                                'Промежуточный метаболизатор': '#FFFF00',
                                                                'Нормальный метаболизатор': '#00B050',
                                                                'Ультрабыстрый метаболизатор': '#00B0F0',
                                                                'Быстрый метаболизатор': '#15EBD2',
                                                                'Неопределенный': '#A6A6A6'},
                                                   subset=['Фенотип'])
 
#nootropic_stimulants_colored.to_excel('H:/Spyder_work/Nootropic_Stimulants.xlsx', engine='openpyxl', index=False)


##Get POR alleles

#por_hap1 = por_df[['Haplotype1']]
#por_hap2 = por_df[['Haplotype2']]
#hap1_str = por_hap1.to_string(index=False, header=False)
#hap2_str = por_hap2.to_string(index=False, header=False)
#hap1_filt = re.findall(r'\W\d+', hap1_str)
#hap1_str = ''.join(hap1_filt)
#hap2_filt = re.findall(r'\W\d+', hap2_str)
#hap2_str = ''.join(hap2_filt)
#por_genotype = hap1_str + '/' + hap2_str


##Create phenotype database for POR

#CYP2D6
cyp2d6_red = ['*2', '*4','*5', '*6','*7', '*8', '*9', '*10', '*11','*12', '*15', '*17', '*18', '*19', '*20', '*21', '*22',
              '*23', '*24', '*38', '*39', '*41']
cyp2d6_yellow = ['*3', '*16', '*27', '*31', '*32', '*33', '*34', '*35', '*36', '*37', '*40']
cyp2d6_green = ['*14', '*28', '*42']
cyp2d6_blue = ['*13']
cyp2d6_normal = ['*1']
cyp2d6_unknown = ['*25', '*26', '*29', '*30', '*43', '*44', '*45', '*46', '*47', '*48']

cyp2d6_red_dict = { red_met : int(0) for red_met in cyp2d6_red }
cyp2d6_yellow_dict = { yel_met : int(0.5) for yel_met in cyp2d6_yellow }
cyp2d6_green_dict = { gre_met : int(0.75) for gre_met in cyp2d6_green }
cyp2d6_blue_dict = { blu_met : int(1.5) for blu_met in cyp2d6_blue }
cyp2d6_normal_dict = { nor_met : int(1) for nor_met in cyp2d6_normal }
cyp2d6_unknown_dict = { unk_met : 'unknown' for unk_met in cyp2d6_unknown }
cyp2d6_por_dict = dict(cyp2d6_normal_dict)
cyp2d6_por_dict.update(cyp2d6_blue_dict)
cyp2d6_por_dict.update(cyp2d6_green_dict)
cyp2d6_por_dict.update(cyp2d6_yellow_dict)
cyp2d6_por_dict.update(cyp2d6_red_dict)
cyp2d6_por_dict.update(cyp2d6_unknown_dict)

def return_score_cyp2d6_por(val):
    for allele, score in cyp2d6_por_dict.items():
        if allele==val:
            return score
    return('unknown')

cyp2d6_score_hap1 = return_score_cyp2d6_por(hap_por_1)
cyp2d6_score_hap2 = return_score_cyp2d6_por(hap_por_2)

def return_phenotype_por(val):
    if val=='unknown':
        return('Неопределенный')
    elif val==0:
        return('Сильное снижение активности')
    elif 1.4>val>=0.6:
        return('Умеренное снижение активности')
    elif 2>val>=1.4:
        return('Незначительное снижение активности')
    elif 2.5>val>=2:
        return('Нормальная активность')
    elif val>=2.5:
        return('Повышение активности')


if cyp2d6_score_hap1 == 'unknown' or cyp2d6_score_hap2 == 'unknown':
    cyp2d6_pheno_por = 'Неопределенный'
else:
    cyp2d6_por_sum = cyp2d6_score_hap1 + cyp2d6_score_hap2
    cyp2d6_pheno_por = return_phenotype_por(cyp2d6_por_sum)

#CYP2C19
cyp2c19_red = ['*2', '*4', '*5', '*6', '*7', '*8', '*9', '*10', '*11','*12', '*15', '*17', '*18', '*19', '*20', '*21', '*22',
              '*23', '*24', '*38', '*39', '*41', '*100']
cyp2c19_yellow = ['*3', '*16', '*27', '*31', '*32', '*33', '*34', '*35', '*36', '*37', '*40', '*101', '*102']
cyp2c19_green = ['*103', '*104', '*105', '*106', '*107']
cyp2c19_blue = ['*13', '*14', '*108', '*109']
cyp2c19_normal = ['*1', '*28', '*42']
cyp2c19_unknown = ['*25', '*26', '*29', '*30', '*43', '*44', '*45', '*46', '*47', '*48']

cyp2c19_red_dict = { red_met : int(0) for red_met in cyp2c19_red }
cyp2c19_yellow_dict = { yel_met : int(0.5) for yel_met in cyp2c19_yellow }
cyp2c19_green_dict = { gre_met : int(0.75) for gre_met in cyp2c19_green }
cyp2c19_blue_dict = { blu_met : int(1.5) for blu_met in cyp2c19_blue }
cyp2c19_normal_dict = { nor_met : int(1) for nor_met in cyp2c19_normal }
cyp2c19_unknown_dict = { unk_met : 'unknown' for unk_met in cyp2c19_unknown }
cyp2c19_por_dict = dict(cyp2c19_normal_dict)
cyp2c19_por_dict.update(cyp2c19_blue_dict)
cyp2c19_por_dict.update(cyp2c19_green_dict)
cyp2c19_por_dict.update(cyp2c19_yellow_dict)
cyp2c19_por_dict.update(cyp2c19_red_dict)
cyp2c19_por_dict.update(cyp2c19_unknown_dict)

def return_score_cyp2c19_por(val):
    for allele, score in cyp2c19_por_dict.items():
        if allele==val:
            return score
    return('unknown')

cyp2c19_score_hap1 = return_score_cyp2c19_por(hap_por_1)
cyp2c19_score_hap2 = return_score_cyp2c19_por(hap_por_1)

if cyp2c19_score_hap1 == 'unknown' or cyp2c19_score_hap2 == 'unknown':
    cyp2c19_pheno_por = 'Неопределенный'
else:
    cyp2c19_por_sum = cyp2c19_score_hap1 + cyp2c19_score_hap2
    cyp2c19_pheno_por = return_phenotype_por(cyp2c19_por_sum)

#CYP1A2
cyp1a2_red = ['*2', '*4', '*5',  '*8', '*9', '*10', '*11','*12', '*15', '*17', '*18', '*19', '*20', '*21', '*22',
              '*23', '*24', '*38', '*39', '*41']
cyp1a2_yellow = ['*3', '*6', '*7', '*16', '*27', '*31', '*32', '*33', '*34', '*35', '*36', '*37', '*40', '*101', '*102', '*105', '*107']
cyp1a2_green = ['*14', '*28', '*42', '*104', '*106', '*109', '*110', '*111']
cyp1a2_blue = ['*13', '*108', '*112', '*113']
cyp1a2_normal = ['*1']
cyp1a2_unknown = ['*25', '*26', '*29', '*30', '*43', '*44', '*45', '*46', '*47', '*48']

cyp1a2_red_dict = { red_met : int(0) for red_met in cyp1a2_red }
cyp1a2_yellow_dict = { yel_met : int(0.5) for yel_met in cyp1a2_yellow }
cyp1a2_green_dict = { gre_met : int(0.75) for gre_met in cyp1a2_green }
cyp1a2_blue_dict = { blu_met : int(1.5) for blu_met in cyp1a2_blue }
cyp1a2_normal_dict = { nor_met : int(1) for nor_met in cyp1a2_normal }
cyp1a2_unknown_dict = { unk_met : 'unknown' for unk_met in cyp1a2_unknown }
cyp1a2_por_dict = dict(cyp1a2_normal_dict)
cyp1a2_por_dict.update(cyp1a2_blue_dict)
cyp1a2_por_dict.update(cyp1a2_green_dict)
cyp1a2_por_dict.update(cyp1a2_yellow_dict)
cyp1a2_por_dict.update(cyp1a2_red_dict)
cyp1a2_por_dict.update(cyp1a2_unknown_dict)

def return_score_cyp1a2_por(val):
    for allele, score in cyp1a2_por_dict.items():
        if allele==val:
            return score
    return('unknown')

cyp1a2_score_hap1 = return_score_cyp1a2_por(hap_por_1)
cyp1a2_score_hap2 = return_score_cyp1a2_por(hap_por_1)
        
if cyp1a2_score_hap1 == 'unknown' or cyp1a2_score_hap2 == 'unknown':
    cyp1a2_pheno_por = 'Неопределенный'
else:
    cyp1a2_por_sum = cyp1a2_score_hap1 + cyp1a2_score_hap2
    cyp1a2_pheno_por = return_phenotype_por(cyp1a2_por_sum)

#CYP3A4_5
cyp3a45_red = ['*2', '*4', '*8', '*9', '*10', '*15', '*18', '*20', '*21', '*22', '*23', '*24', '*38', '*39', '*41']
cyp3a45_yellow = ['*3', '*5', '*6', '*7', '*12', '*16', '*17', '*27', '*40']
cyp3a45_green = ['*11', '*14', '*19', '*31', '*32', '*33', '*34', '*35']
cyp3a45_blue = ['*13']
cyp3a45_normal = ['*1', '*28', '*36', '*37', '*42']
cyp3a45_unknown = ['*25', '*26', '*29', '*30', '*43', '*44', '*45', '*46', '*47', '*48']

cyp3a45_red_dict = { red_met : int(0) for red_met in cyp3a45_red }
cyp3a45_yellow_dict = { yel_met : int(0.5) for yel_met in cyp3a45_yellow }
cyp3a45_green_dict = { gre_met : int(0.75) for gre_met in cyp3a45_green }
cyp3a45_blue_dict = { blu_met : int(1.5) for blu_met in cyp3a45_blue }
cyp3a45_normal_dict = { nor_met : int(1) for nor_met in cyp3a45_normal }
cyp3a45_unknown_dict = { unk_met : 'unknown' for unk_met in cyp3a45_unknown }
cyp3a45_por_dict = dict(cyp3a45_normal_dict)
cyp3a45_por_dict.update(cyp3a45_blue_dict)
cyp3a45_por_dict.update(cyp3a45_green_dict)
cyp3a45_por_dict.update(cyp3a45_yellow_dict)
cyp3a45_por_dict.update(cyp3a45_red_dict)
cyp3a45_por_dict.update(cyp3a45_unknown_dict)

def return_score_cyp3a45_por(val):
    for allele, score in cyp3a45_por_dict.items():
        if allele==val:
            return score
    return('unknown')

cyp3a45_score_hap1 = return_score_cyp3a45_por(hap_por_1)
cyp3a45_score_hap2 = return_score_cyp3a45_por(hap_por_1)

if cyp3a45_score_hap1 == 'unknown' or cyp3a45_score_hap2 == 'unknown':
    cyp3a45_pheno_por = 'Неопределенный'
else:
    cyp3a45_por_sum = cyp3a45_score_hap1 + cyp3a45_score_hap2
    cyp3a45_pheno_por = return_phenotype_por(cyp3a45_por_sum)

#Write POR to Genotype
por_dict = {'Ген': ['POR'], 'Аллели': [por_genotype], 'Фенотип': ['Вкладка POR_impact']}
por_df = pd.DataFrame(por_dict)
result_df = result_df.append(por_df, ignore_index=True)

#POR_impact
impact_dict = {'Ген': ["CYP2D6", "CYP2C19", "CYP1A2", "CYP3A4", "CYP3A5"], 
               'Фенотип': [cyp2d6_pheno_por, cyp2c19_pheno_por, cyp1a2_pheno_por, cyp3a45_pheno_por, cyp3a45_pheno_por]}
por_impact = pd.DataFrame(impact_dict)

#Apply color for POR_impact
por_impact_colored = por_impact.style.apply(color_col,
                                          pattern_map={'Сильное снижение активности': '#FF0000',
                                                       'Умеренное снижение активности': '#FFFF00',
                                                       'Незначительное снижение активности': '#7AE06C',
                                                       'Нормальная активность': '#00B050',
                                                       'Повышение активности': '#00B0F0',
                                                       'Неопределенный': '#A6A6A6'},
                                          subset=['Фенотип'])

#Apply color for Genotype
result_df_colored = result_df.style.apply(color_col,
                                          pattern_map={'Медленный метаболизатор': '#FF0000',
                                                       'Промежуточный метаболизатор': '#FFFF00',
                                                       'Нормальный метаболизатор': '#00B050',
                                                       'Ультрабыстрый метаболизатор': '#00B0F0',
                                                       'Быстрый метаболизатор': '#15EBD2',
                                                       'Неопределенный': '#A6A6A6'},
                                          subset=['Фенотип'])

#First page info
input_fname = os.path.basename(sys.argv[1])
input_ff = {'Образец': [input_fname]}
info_pp = pd.DataFrame(input_ff)


##Combine all results
shutil.copy(script_path + 'Risk_alleles_Legend.xlsx', sys.argv[2])
wb = load_workbook(sys.argv[2])
writer = pd.ExcelWriter(sys.argv[2], engine = 'openpyxl')
writer.book = wb
info_pp.to_excel(writer, index=False, sheet_name = 'Info')
result_df_colored.to_excel(writer, index=False, sheet_name = 'Genotype')
neuroleptics_colored.to_excel(writer, index=False, sheet_name = 'Neuroleptics')
antidepressants_colored.to_excel(writer, index=False, sheet_name = 'Antidepressants')
anticonvulsants_colored.to_excel(writer, index=False, sheet_name = 'Anticonvulsants_Normotimics')
tranquilizers_colored.to_excel(writer, index=False, sheet_name = 'Tranquilizers_Sedatives')
nootropic_stimulants_colored.to_excel(writer, index=False, sheet_name = 'Nootropic_Stimulants')
por_impact_colored.to_excel(writer, index=False, sheet_name = 'POR_impact')
wb.move_sheet("Risk_alleles", (len(wb.sheetnames)-1))
wb.move_sheet("Legend", (1))
writer.save()
#writer.close()


##Write JSON
#print(por_impact.to_json())
#result_dt_to_html = result_df_colored.render()
#print(result_df_colored.to_html())


# Excel to PDF
#from win32com import client
  
# Open Microsoft Excel
#excel = client.Dispatch("Excel.Application")
  
# Read Excel File
#sheets = excel.Workbooks.Open('C:/Users/acer/Desktop/Patient_result.xlsx')
#work_sheets = sheets.Worksheets['Neuroleptics']

# Convert into PDF File
#work_sheets.ExportAsFixedFormat(0, 'C:/Users/acer/Desktop/result.pdf')

# Auto-adjust columns width
#writer_adj = pd.ExcelWriter('C:/Users/acer/Desktop/Patient_result.xlsx', engine = 'xlsxwriter')

#result_df.to_excel(writer_adj, sheet_name='Genotype', index=False, na_rep='NaN')
#for column in result_df:
#    column_length = max(result_df[column].astype(str).map(len).max(), len(column))
#    col_idx = result_df.columns.get_loc(column)
#    writer_adj.sheets['Genotype'].set_column(col_idx, col_idx, column_length)
